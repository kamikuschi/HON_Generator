from openpyxl import load_workbook

# Pfad zur Excel-Datei (du musst den Pfad zu deiner Datei anpassen)
excel_path = '/mnt/data/your_excel_file.xlsx'

# Excel-Datei laden
workbook = load_workbook(excel_path)
sheet = workbook.active

# Variablen für die Daten
dates = []
start_times = []
end_times = []
work_times = []

# Automatische Erkennung der ersten Datenzeile
data_start_row = 1
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=2, max_col=2, values_only=True):
    data_start_row += 1
    if row[0] is not None:
        break

# Zeilen durchlaufen ab der erkannten Datenzeile
for row in sheet.iter_rows(min_row=data_start_row, min_col=2, max_col=8, values_only=True):
    date, weekday, start_time, end_time, work_time, _, is_settled = row
    
    # Überprüfen, ob das Feld in Spalte H leer ist
    if is_settled is None:
        # Daten zu den Listen hinzufügen
        dates.append(date)
        start_times.append(start_time)
        end_times.append(end_time)
        work_times.append(work_time)

# Ausgabe der Variablen
print("Dates:", dates)
print("Start Times:", start_times)
print("End Times:", end_times)
print("Work Times:", work_times)
